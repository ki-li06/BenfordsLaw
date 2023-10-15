from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([3, 2, 1])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

print("test_Openpyxl.py loaded successfully!")

# Save the file
wb.save("sample.xlsx")