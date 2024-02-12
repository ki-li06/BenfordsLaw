import math
from openpyxl import Workbook
from BenfordsLaw import prob_d_i

wb = Workbook()

ws = wb.active

ws.cell(column=1, row=1).value = "d"

for i in range(1, 5):
    ws.cell(column=i+1, row=1).value = "P(D_" + str(i) + "=d)"

for d in range(0, 10):
    ws.cell(column=1, row = d+1+1).value = d
    for i in range(1, 5):
        value = prob_d_i(d, i)*100
        print("value for d=", d, " and i=", i, ": ", value)
        ws.cell(column=i+1, row= d+1+1).value = value

wb.save("ExcelSheets/Output/P(D_i=d).xlsx")
