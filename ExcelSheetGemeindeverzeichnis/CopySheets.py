import os
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from Main import get_orig_sheet, get_new_sheet, save_workbook
import time

# Get workbook sheet object
sheet_orig = get_orig_sheet()
sheet_new = get_new_sheet()
 
# Cell objects also have a row, column, 
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or 
# column integer is 1, not 0.
 

row = 7
cell_row = sheet_orig[row]
cell_satzart = cell_row[0]
#print("satzart:", cell_satzart.value)
#print("coordinate:", cell_satzart.coordinate)
elements = 0

while cell_satzart.value != None:
#while elements < 10:
    SATZART_GEMEINDE = 60
    #print("satzart = " + str(cell_satzart.value) + " (type=" + str(type(cell_satzart.value)) + ")")
    if int(cell_satzart.value) == SATZART_GEMEINDE:
        values = expected_freqs = {i: None for i in range(1, 16)}
        for i in range(1, 16):
            values[i] = cell_row[i].value
        #cell_bev_ins = cell_row[9]
        for i in range(2, 7):
            values[i] = str(values[i])
        ars = values[2] + values[3] + values[4] + values[5] + values[6]
        name = values[7]
        fläche = values[8]
        bev = values[9], values[10], values[11], values[12]
        plz = values[13]
        if values[14] == None:
            koord = None, None
        else:            
            koord = float(str(values[14]).replace(",", ".")), float(str(values[15]).replace(",", "."))
        #print("row=" +  str(row) + " => " + str(ars) + " " + str(name) + " (" + str(fläche) + " km²) " + str(bev) + " " + str(plz) + " " + str(koord))
        #print("types:", type(ars), type(name), type(fläche), [type(bev[i]) for i in range(0, len(bev))], type(plz), type(koord[0]), type(koord[1]))
        print("row=" + str(row) + " => " + str(name))

        sheet_new.cell(column=1, row = elements + 7).value = ars
        sheet_new.cell(column=2, row = elements + 7).value = name
        sheet_new.cell(column=3, row = elements + 7).value = fläche
        sheet_new.cell(column=4, row = elements + 7).value = bev[0]
        sheet_new.cell(column=5, row = elements + 7).value = bev[1]
        sheet_new.cell(column=6, row = elements + 7).value = bev[2]
        sheet_new.cell(column=7, row = elements + 7).value = bev[3]
        sheet_new.cell(column=8, row = elements + 7).value = plz
        sheet_new.cell(column=9, row = elements + 7).value = koord[0]
        sheet_new.cell(column=9, row = elements + 7).alignment = openpyxl.styles.Alignment(horizontal='right')
        sheet_new.cell(column=10, row = elements + 7).value = koord[1]
        sheet_new.cell(column=10, row = elements + 7).alignment = openpyxl.styles.Alignment(horizontal='right')

        elements = elements + 1
    #time.sleep(3)  # add a 20 millisecond delay
    row = row + 1
    #print("row:", row)
    cell_row = sheet_orig[row][:20]
    cell_satzart = cell_row[0]

print("elements:", elements)
#overall there are 10981 elements

# Save the workbook
save_workbook()

print("finished copying sheets")