import openpyxl
from openpyxl import load_workbook

wb_orig = load_workbook(filename = 'ExcelSheets/Input/Gemeindeverzeichnis - Vorlage.xlsx')

sheet_orig = wb_orig.worksheets[1]
sheet_new = wb_orig["GemeindeverzeichnisGekürzt"]
 
row = 7
cell_row = sheet_orig[row]
cell_satzart = cell_row[0]
elements = 0

while cell_satzart.value != None:
    SATZART_GEMEINDE = 60
    if int(cell_satzart.value) == SATZART_GEMEINDE:
        values = expected_freqs = {i: None for i in range(1, 16)}
        for i in range(1, 16):
            if(i in range(2, 7)):
                values[i] = str(cell_row[i].value)
            else:
                values[i] = cell_row[i].value

        ars = values[2] + values[3] + values[4] + values[5] + values[6]
        name = values[7]
        fläche = values[8]
        bev = values[9], values[10], values[11], values[12]
        plz = values[13]
        if values[14] == None:
            koord = None, None
        else:            
            koord = float(str(values[14]).replace(",", ".")), float(str(values[15]).replace(",", "."))
        print("row=" + str(row) + " => " + str(name))

        sheet_new.cell(column=1, row = elements + 7).value = ars
        sheet_new.cell(column=2, row = elements + 7).value = name
        sheet_new.cell(column=3, row = elements + 7).value = fläche
        sheet_new.cell(column=4, row = elements + 7).value = bev[0]
        sheet_new.cell(column=5, row = elements + 7).value = bev[1]
        sheet_new.cell(column=6, row = elements + 7).value = bev[2]
        sheet_new.cell(column=7, row = elements + 7).value = bev[3]
        sheet_new.cell(column=8, row = elements + 7).value = plz
        sheet_new.cell(column=9, row = elements + 7).value = koord[0]
        sheet_new.cell(column=9, row = elements + 7).alignment = openpyxl.styles.Alignment(horizontal='right')
        sheet_new.cell(column=10, row = elements + 7).value = koord[1]
        sheet_new.cell(column=10, row = elements + 7).alignment = openpyxl.styles.Alignment(horizontal='right')

        elements = elements + 1
    row = row + 1
    cell_row = sheet_orig[row][:20]
    cell_satzart = cell_row[0]

print("elements:", elements)
#overall there should be 10981 elements

wb_orig.save("ExcelSheets/Input/Gemeindeverzeichnis - Gekürzt.xlsx")

print("finished copying sheets")