import openpyxl
from openpyxl import load_workbook

print("loading workbook...")
wb_obj = load_workbook("ExcelSheets/Input/Gemeindeverzeichnis - Gekürzt.xlsx")
print("-> done")
sheet = wb_obj["GemeindeverzeichnisGekürzt"]

def get_column_data(column: str):
    encoder = {"ars": 1, "name": 2, "fläche": 3, "bev0": 4, "bev1": 5, "bev2": 6, "bev3": 7, "plz": 8, "koord0": 9, "koord1": 10}
    if type(column) == str:
        column_i = encoder[column]
    else:
        column_i = column
        column = list(encoder.keys())[list(encoder.values()).index(column)]
    row = 7
    data = []
    NoneElements = 0
    ZeroElements = 0
    while True:
        cell_ars = sheet.cell(row,1)
        cell = sheet.cell(row, column_i)
        if cell_ars.value == None:
            break
        elif cell.value == None:
            NoneElements += 1
            pass
        elif cell.value == 0:
            ZeroElements += 1
            pass
        else:
            data.append(cell.value)
        row += 1
    if NoneElements > 0 or ZeroElements > 0:
        print("column", column,"contains invalid elements --- NoneElements:", NoneElements, " ZeroElements:", ZeroElements)
    return data

list = get_column_data(3)
print(len(list), " elements in list")

