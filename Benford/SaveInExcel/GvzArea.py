import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import sys
sys.path.append('ExcelSheetGemeindeverzeichnis')
from GetData import get_column_data
sys.path.append('Benford')
from AnalysisBenfordsLaw import benfords_law_on_dataset, get_biggest_difference
from Display import plot_dataframe, print_as_dataframe

mylist = get_column_data("fläche")
print("first 10 elements:", str(mylist[:10]))
benford = benfords_law_on_dataset(mylist)

print_as_dataframe(benford)

highest_dif = get_biggest_difference(benford)
print("highest difference:", highest_dif)

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers for the table
worksheet['A1'] = 'i'
worksheet['B1'] = 'h(D_1=i)'

# Loop through the range of values for i and write the corresponding Fibonacci(i) value in the second column
for i in range(1, 10):
    worksheet.cell(row=i+1, column=1).value = i
    worksheet.cell(row=i+1, column=2).value = benford["actual"][i]

workbook.save('0_Excels\Diagrams\BenfordOnFib.xlsx')


