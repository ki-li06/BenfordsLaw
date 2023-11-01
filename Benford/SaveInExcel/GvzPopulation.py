import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

import sys
sys.path.append('ExcelSheetGemeindeverzeichnis')
from GetData import get_column_data
sys.path.append('Benford')
from AnalysisBenfordsLaw import benfords_law_on_dataset, get_biggest_difference
from Display import plot_dataframe, print_as_dataframe

def analyse_bev_column(column_name, meaning: str):
    if(type(column_name) == str):
        mylist = get_column_data(column_name)
        print("first 10 elements:", str(mylist[:10]))
    else :
        mylist = []
        for i in range(0, len(column_name)):
            column = column_name[i]
            list_of_column = get_column_data(column)
            for j in range(0, len(list_of_column)):
                list_of_column[j] = int(list_of_column[j])
                element = list_of_column[j]
                mylist.append(element)
        print("first 10 elements:", str(mylist[:10]))
    benford = benfords_law_on_dataset(mylist)
    print("-"*50)
    print(meaning)
    print_as_dataframe(benford)
    highest_dif = get_biggest_difference(benford)
    print("highest difference:", highest_dif)
    return (benford, highest_dif)

benford_bev_ges = analyse_bev_column("bev0", "Bevölkerung gesamt")
benford_bev_m = analyse_bev_column("bev1", "Bevölkerung männlich")
benford_bev_w = analyse_bev_column("bev2", "Bevölkerung weiblich")
benford_bev_dichte = analyse_bev_column("bev3", "Bevölkerungsdichte")
benford_bev_mix = analyse_bev_column(["bev0", "bev1", "bev2", "bev3"], "Bevölkerung mix")

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers for the table
worksheet['A1'] = 'd'
worksheet['B1'] = 'P(D₁=d)'
worksheet['C1'] = 'Gesamtbevölkerung'
worksheet['D1'] = 'Männliche Bevölkerung'
worksheet['E1'] = 'Weibliche Bevölkerung'
worksheet['F1'] = 'Bevölkerungsdichte'
worksheet['G1'] = 'Mix-Bevölkerung'

# Loop through the range of values for i and write the corresponding Fibonacci(i) value in the second column
for i in range(1, 10):
    worksheet.cell(row=i+1, column=1).value = i
    worksheet.cell(row=i+1, column=2).value = 100*benford_bev_ges[0]["theoretic"][i]
    worksheet.cell(row=i+1, column=3).value = 100*benford_bev_ges[0]["actual"][i]
    worksheet.cell(row=i+1, column=4).value = 100*benford_bev_m[0]["actual"][i]
    worksheet.cell(row=i+1, column=5).value = 100*benford_bev_w[0]["actual"][i]
    worksheet.cell(row=i+1, column=6).value = 100*benford_bev_dichte[0]["actual"][i]
    worksheet.cell(row=i+1, column=7).value = 100*benford_bev_mix[0]["actual"][i]

worksheet.cell(row=12, column=1).value = "Δ"
worksheet.cell(row=12, column=3).value = 100*benford_bev_ges[1]
worksheet.cell(row=12, column=4).value = 100*benford_bev_m[1]
worksheet.cell(row=12, column=5).value = 100*benford_bev_w[1]
worksheet.cell(row=12, column=6).value = 100*benford_bev_dichte[1]
worksheet.cell(row=12, column=7).value = 100*benford_bev_mix[1]

workbook.save('0_Excels\Diagrams\GVZ\Bevoelkerung.xlsx')


