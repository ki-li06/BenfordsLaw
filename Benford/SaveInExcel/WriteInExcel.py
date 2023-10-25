from BenfordSequences import fibonacci
import openpyxl

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers for the table
worksheet['A1'] = 'i'
worksheet['B1'] = 'Fibonacci(i)'

n = 100


# Loop through the range of values for i and write the corresponding Fibonacci(i) value in the second column
for i in range(1, 21):
    worksheet.cell(row=i+1, column=1).value = i
    worksheet.cell(row=i+1, column=2).value = fibonacci(i)


workbook.save('./0_Excels/Diagrams/BenfordOnFib.xlsx')
