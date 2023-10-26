import openpyxl

import sys
sys.path.append('Benford')
from BenfordSequences import fibonacci
from AnalysisBenfordsLaw import benfords_law_on_dataset
from Display import plot_dataframe

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers for the table
worksheet['A1'] = 'i'
worksheet['B1'] = 'fibonacci(i)'

n = 100

fib = fibonacci(n)
print("collatz sequence from " + str(n) + " -> " + str(fib) + " (length: " + str(len(fib)) + ")")

benford = benfords_law_on_dataset(fib)

highest_difference = 0

# Loop through the range of values for i and write the corresponding Fibonacci(i) value in the second column
for i in range(1, 10):
    worksheet.cell(row=i+1, column=1).value = i
    value = round(benford["actual"][i]*100, 1)
    worksheet.cell(row=i+1, column=2).value = value
    print("i=", i, " -> ", value, "%")
    difference = round(abs(benford["actual"][i] - benford["expected"][i])*100,2)
    if(difference > highest_difference):
        highest_difference = difference
        #print("new highest difference: ", highest_difference, "%"+' for i=', i)

print("highest difference: ", highest_difference, "%")

print("-"*50)

n = 1000
fib = fibonacci(n)
print("collatz sequence from " + str(n) + " -> " + str(fib[10:]) + " (length: " + str(len(fib)) + ")")

benford = benfords_law_on_dataset(fib)

highest_difference = 0

for i in range(1, 10):
    value = round(benford["actual"][i]*100, 1)
    worksheet.cell(row=i+1, column=3).value = value
    print("i=", i, " -> ", value, "%")
    difference = round(abs(benford["actual"][i] - benford["expected"][i])*100,2)
    if(difference > highest_difference):
        highest_difference = difference
        #print("new highest difference: ", highest_difference, "%"+' for i=', i)

print("highest difference: ", highest_difference, "%")

workbook.save('0_Excels\Diagrams\BenfordOnFib.xlsx')


#plot_dataframe(benford, x_label="First digit", y_label="Frequency", title="Benford's law on the Fibonacci sequence (" + str(len(fib)) + " elements)")