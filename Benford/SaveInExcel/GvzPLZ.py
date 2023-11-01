import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import sys
sys.path.append('ExcelSheetGemeindeverzeichnis')
from GetData import get_column_data
sys.path.append('Benford')
from AnalysisBenfordsLaw import benfords_law_on_dataset, get_biggest_difference
from Display import plot_dataframe, print_as_dataframe

mylist = get_column_data("plz")
print("first 10 elements:", str(mylist[:10]))
benford = benfords_law_on_dataset(mylist)

print_as_dataframe(benford)

highest_dif = get_biggest_difference(benford)
print("highest difference:", highest_dif)

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers for the table
worksheet['A1'] = 'd'
worksheet['B1'] = 'P(D₁=d)'
worksheet['C1'] = 'PLZen'

# Loop through the range of values for i and write the corresponding Fibonacci(i) value in the second column
for i in range(1, 10):
    worksheet.cell(row=i+1, column=1).value = i
    worksheet.cell(row=i+1, column=2).value = 100*benford["theoretic"][i]
    worksheet.cell(row=i+1, column=3).value = 100*benford["actual"][i]

worksheet.cell(row=12, column=1).value = "Δ"
worksheet.cell(row=12, column=3).value = 100*highest_dif



workbook.save('0_Excels\Diagrams\GVZ\PLZen.xlsx')


