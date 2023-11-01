import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

import sys
sys.path.append('ExcelSheetGemeindeverzeichnis')
from GetData import get_column_data
sys.path.append('Benford')
from AnalysisBenfordsLaw import benfords_law_on_dataset, get_biggest_difference
from Display import plot_dataframe, print_as_dataframe

def analyse_bev_column(column_name, meaning: str):
    mylist = get_column_data(column_name)
    print("first 10 elements:", str(mylist[:10]))
    benford = benfords_law_on_dataset(mylist)
    print("-"*50)
    print(meaning)
    print_as_dataframe(benford)
    highest_dif = get_biggest_difference(benford)
    print("highest difference:", highest_dif)
    return (benford, highest_dif)

benford_koord_länge = analyse_bev_column("koord0", "Längengrad")
benford_koord_breite = analyse_bev_column("koord1", "Breitengrad")

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers for the table
worksheet['A1'] = 'd'
worksheet['B1'] = 'P(D₁=d)'
worksheet['C1'] = 'Längengrad'
worksheet['D1'] = 'Breitengrad'

# Loop through the range of values for i and write the corresponding Fibonacci(i) value in the second column
for i in range(1, 10):
    worksheet.cell(row=i+1, column=1).value = i
    worksheet.cell(row=i+1, column=2).value = 100*benford_koord_länge[0]["theoretic"][i]
    worksheet.cell(row=i+1, column=3).value = 100*benford_koord_länge[0]["actual"][i]
    worksheet.cell(row=i+1, column=4).value = 100*benford_koord_breite[0]["actual"][i]

worksheet.cell(row=12, column=1).value = "Δ"
worksheet.cell(row=12, column=3).value = 100*benford_koord_länge[1]
worksheet.cell(row=12, column=4).value = 100*benford_koord_breite[1]

langengrade = get_column_data("koord0")
breitengrad = get_column_data("koord1")

worksheet.cell(row=15, column=1).value = "Maximal"
worksheet.cell(row=15, column=2).value = max(langengrade)
worksheet.cell(row=15, column=3).value = max(breitengrad)
worksheet.cell(row=16, column=1).value = "Minimal"
worksheet.cell(row=16, column=2).value = min(langengrade)
worksheet.cell(row=16, column=3).value = min(breitengrad)

workbook.save('0_Excels\Diagrams\GVZ\Koordinaten.xlsx')


